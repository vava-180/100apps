# -*- coding: utf-8 -*-
"""
複数Excel→1ブック集約ツール（Day 039）

フォルダの中にある複数のExcel（.xlsx）の各シートを、1つのExcelブックにまとめます。
別々のファイルに分かれた表を、1ファイルにシートとして集めたいときに使います。

しくみ:
  1) フォルダ内の .xlsx を集める（名前順）
  2) 各ファイルの各シートを読み、1つの出力ブックに新しいシートとして追加する
  3) シート名はファイル名をもとに付ける（Excelの31文字制限・重複は自動で調整）
写すのは「セルの値」だけです（列幅・色・罫線・結合セルなどの書式は写りません）。
数式の結果は、一度Excelで開いて保存したファイルなら反映されます
（保存していない数式は空になることがあります）。

外部ライブラリは openpyxl を使います（Excelをローカルで読み書きするためのもの・ネット通信なし）。
  インストール（1回だけ）:  python -m pip install openpyxl
元のExcelは書き換えません（読み取り＋新しいブックの書き出しだけ）。
使い方: ターミナルで  python merge_excel.py   と実行するか、ファイルをダブルクリック。
"""

import re
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(errors="replace")
except Exception:
    pass

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl が見つかりません。先に  python -m pip install openpyxl  を実行してください。")
    try:
        input("\nEnterキーで終了します…")
    except EOFError:
        pass
    sys.exit(1)

# Excelのシート名で使えない文字と、長さの上限
INVALID_SHEET_CHARS = r"\/?*[]:"
MAX_SHEET_LEN = 31
# このツールが作る出力ファイル名（再実行時に過去の出力を集め直さないため）
OUTPUT_BASENAME = "集約結果"
OUTPUT_RE = re.compile(r"^集約結果(\(\d+\))?\.xlsx$", re.IGNORECASE)


# ===== 中心ロジック（テストしやすい純粋関数）=====

def sanitize_sheet_title(name: str) -> str:
    """Excelのシート名に使えない文字を _ にし、31文字までに収める。空なら 'シート'。"""
    cleaned = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in name)
    cleaned = cleaned.strip().strip("'")        # 先頭末尾の空白とアポストロフィは避ける
    cleaned = cleaned[:MAX_SHEET_LEN]
    cleaned = cleaned.strip().strip("'")        # 切り詰めた結果、末尾が ' などになるのを整える
    if not cleaned:
        cleaned = "シート"
    return cleaned


def unique_sheet_title(existing, base: str) -> str:
    """既にあるシート名と重ならない名前にする。31文字を超えないよう連番分を切り詰める。

    Excelのシート名は大文字小文字を区別しないので、比較は小文字にそろえて行う。
    """
    existing_lower = {x.lower() for x in existing}
    title = sanitize_sheet_title(base)
    if title.lower() not in existing_lower:
        return title
    i = 2
    while True:
        suffix = f"({i})"
        head = title[:MAX_SHEET_LEN - len(suffix)]  # 連番を付けても31字に収まるよう本体を削る
        cand = head + suffix
        if cand.lower() not in existing_lower:
            return cand
        i += 1


def make_sheet_base(file_stem: str, sheet_name: str, single_sheet: bool) -> str:
    """出力シート名のもと。ファイルが1シートならファイル名、複数なら『ファイル名_シート名』。"""
    if single_sheet:
        return file_stem
    return f"{file_stem}_{sheet_name}"


# ===== ファイル集め＝集約処理 =====

def is_temp_excel(name: str) -> bool:
    """Excelが開いている間だけ作る一時ファイル（~$ で始まる）かどうか。"""
    return name.startswith("~$")


def collect_xlsx(folder: Path, exclude: set):
    """フォルダ直下の .xlsx を集める（名前順）。

    除くもの: Excelの一時ファイル(~$)、exclude（絶対パス集合）、
    そしてこのツール自身の出力（集約結果.xlsx / 集約結果(n).xlsx）。
    過去の集約結果を再び集めて重複・肥大化させないため。
    """
    files = []
    for p in folder.iterdir():
        if not p.is_file() or p.suffix.lower() != ".xlsx":
            continue
        if is_temp_excel(p.name) or p.resolve() in exclude:
            continue
        if OUTPUT_RE.match(p.name):  # 過去に出した集約結果は入力に含めない
            continue
        files.append(p)
    return sorted(files, key=lambda x: x.name.lower())


def copy_sheet_values(src_ws, dst_ws):
    """元シートのセルの値を、新しいシートへ1行ずつ写す。値の数を返す。"""
    count = 0
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(row)
        count += sum(1 for v in row if v is not None)
    return count


def merge_files(files, out_wb):
    """集めたファイルの各シートを out_wb に追加する。(追加シート数, 失敗一覧) を返す。

    1ファイル・1シートが壊れていても、そこだけ失敗として記録し、残りは続ける。
    """
    added = 0
    failed = []
    for path in files:
        try:
            # data_only=True … 数式ではなく「計算結果（保存済みの値）」を読む
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as e:  # ファイル自体が開けない（壊れているなど）
            failed.append((path.name, str(e)))
            continue
        try:
            sheetnames = wb.sheetnames
            single = len(sheetnames) == 1
            for sname in sheetnames:
                dst = None
                try:
                    src = wb[sname]
                    base = make_sheet_base(path.stem, sname, single)
                    title = unique_sheet_title(set(out_wb.sheetnames), base)
                    dst = out_wb.create_sheet(title=title)
                    copy_sheet_values(src, dst)
                    added += 1
                except Exception as e:  # 読み取り途中の破損など。そのシートだけ飛ばす
                    if dst is not None and dst in out_wb.worksheets:
                        out_wb.remove(dst)  # 作りかけのシートを残さない
                    failed.append((f"{path.name}[{sname}]", str(e)))
        finally:
            wb.close()  # read_only で開いたファイルを確実に閉じる
    return added, failed


def unique_path(path: Path):
    """同名ファイルがあれば連番を付ける。空き名が無ければ None。"""
    if not path.exists():
        return path
    for i in range(2, 1000):
        cand = path.with_name(f"{path.stem}({i}){path.suffix}")
        if not cand.exists():
            return cand
    return None


# ===== 画面とのやり取り =====

def ask(prompt: str, default: str = "") -> str:
    suffix = f"（未入力なら {default}）" if default else ""
    try:
        value = input(f"{prompt}{suffix}: ").strip()
    except EOFError:
        return default
    return value if value else default


def pause_and_exit():
    try:
        input("\nEnterキーで終了します…")
    except EOFError:
        pass


def main():
    print("=" * 48)
    print(" 複数Excel→1ブック集約ツール（Day 039）")
    print("=" * 48)

    here = str(Path(__file__).parent)
    folder = Path(ask("Excel(.xlsx)が入ったフォルダ", here).strip().strip('"').strip("'"))
    if not folder.is_dir():
        print(f"⚠ フォルダが見つかりません: {folder}")
        pause_and_exit()
        return

    # これから作る出力ファイル名を先に決め、集める対象から外す
    out_path = unique_path(folder / f"{OUTPUT_BASENAME}.xlsx")
    if out_path is None:
        print("⚠ 集約結果ファイルが多すぎます。古い『集約結果*.xlsx』を移動してください。")
        pause_and_exit()
        return

    files = collect_xlsx(folder, {out_path.resolve()})
    if not files:
        print("フォルダの中に .xlsx が見つかりませんでした。")
        pause_and_exit()
        return

    print(f"\n{len(files)} 個のExcelを1つのブックにまとめます:")
    for p in files:
        print(f"   - {p.name}")
    if ask("\n実行しますか？ (yes/no)", "yes").lower() not in ("yes", "y"):
        print("実行しませんでした。")
        pause_and_exit()
        return

    out_wb = Workbook()
    # 最初から入っている空シートは、あとで中身を入れた後に消す
    default_sheet = out_wb.active

    added, failed = merge_files(files, out_wb)

    if added == 0:
        print("\n集約できるシートがありませんでした。")
        pause_and_exit()
        return

    # 自動でできた空の初期シートを削除（実データのシートが1つ以上ある場合）
    if default_sheet.title in out_wb.sheetnames and len(out_wb.sheetnames) > 1:
        out_wb.remove(default_sheet)

    try:
        out_wb.save(out_path)
    except OSError as e:
        print(f"⚠ 書き出しに失敗しました: {e}")
        pause_and_exit()
        return

    print(f"\n✅ {added} シートを集約しました（値のみコピー・書式や結合セルは写しません）:\n   {out_path}")
    if failed:
        print("⚠ 読めずに飛ばしたファイル:")
        for name, why in failed:
            print(f"   - {name}: {why}")
    pause_and_exit()


if __name__ == "__main__":
    main()
