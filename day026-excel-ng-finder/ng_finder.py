# -*- coding: utf-8 -*-
"""
Excel「NG箇所」検出＆一覧化ツール（Day 026）

Excelで作ったチェックリストなどから、指定した文字（例: NG）を含むセルを
全シートまとめて探し出し、新しいシート「検出結果」に一覧化します。
一覧の各行には「元のセルへ飛ぶリンク」を付けるので、クリックでその箇所へジャンプできます。

しくみ:
  1) ブック内の全シート・全セルを順に見る
  2) 指定文字に当てはまるセルを集める（部分一致／完全一致・大文字小文字の有無を選べる）
  3) 新しいシートに「シート名・セル番地・値・リンク」を一覧化
元ファイルは上書きせず、別名（〇〇_検出結果.xlsx）で保存します（安全のため）。

外部ライブラリ openpyxl を使います（Excelをローカルで読み書きするだけ・ネット通信なし）。
最初の1回だけ:  pip install openpyxl
使い方: ターミナルで  python ng_finder.py   と実行するか、ファイルをダブルクリック。
"""

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(errors="replace")
except Exception:
    pass

# openpyxl が入っていない場合は、やさしく案内して終了する
try:
    import openpyxl
    from openpyxl.utils import get_column_letter  # noqa: F401  (将来用・存在確認も兼ねる)
except ImportError:
    print("⚠ このツールには『openpyxl』が必要です。次のコマンドで1回だけ入れてください:")
    print("    pip install openpyxl")
    try:
        input("\nEnterキーで終了します…")
    except EOFError:
        pass
    sys.exit(1)


# ===== 判定の中心ロジック（openpyxl不要・テストしやすい純粋関数）=====

def text_matches(value, keyword: str, mode: str, case_sensitive: bool) -> bool:
    """セルの値が、探したい文字に当てはまるか判定する。

    value          : セルの値（None・数値・文字列など何でも来る）
    keyword        : 探す文字（例: "NG"）
    mode           : "contains"（部分一致）か "exact"（完全一致）
    case_sensitive : True なら大文字小文字を区別する
    """
    if value is None:
        return False
    text = str(value)
    key = keyword
    if not case_sensitive:
        text = text.lower()
        key = key.lower()
    if mode == "exact":
        return text.strip() == key.strip()
    return key in text  # 既定は部分一致


def build_link_formula(sheet_name: str, coord: str, display: str) -> str:
    """同じブック内の『そのセル』へ飛ぶ HYPERLINK 式を組み立てる。

    シート名に含まれる ' は Excel の決まりで '' に2重化し、値の " も "" にして壊れないようにする。
    """
    safe_sheet = sheet_name.replace("'", "''")
    # 先頭の # が「同じブック内へのリンク」の合図。数式は " で囲むので、中の " は "" にして壊さない
    target = f"#'{safe_sheet}'!{coord}".replace('"', '""')
    safe_disp = str(display).replace('"', '""')
    return f'=HYPERLINK("{target}","{safe_disp}")'


def unique_sheet_title(existing: list, base: str = "検出結果") -> str:
    """既にあるシート名と重ならないタイトルを作る（検出結果, 検出結果(2), …）。"""
    if base not in existing:
        return base
    i = 2
    while f"{base}({i})" in existing:
        i += 1
    return f"{base}({i})"


def unique_path(path: Path) -> Path:
    """同名ファイルがあれば連番を付けて、上書きを避けたパスを返す。"""
    if not path.exists():
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    i = 2
    while True:
        cand = parent / f"{stem}({i}){suffix}"
        if not cand.exists():
            return cand
        i += 1


# ===== Excelの読み取りと書き出し =====

def find_matches(wb, keyword: str, mode: str, case_sensitive: bool, skip_titles: set) -> list:
    """ブック内の全シートを走査し、当てはまるセルを集める。

    戻り値: [(シート名, セル番地, セルの値), ...]
    skip_titles に入っているシートは対象外（過去に作った「検出結果」などを除くため）。
    """
    hits = []
    for ws in wb.worksheets:
        if ws.title in skip_titles:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if text_matches(cell.value, keyword, mode, case_sensitive):
                    hits.append((ws.title, cell.coordinate, cell.value))
    return hits


def write_result_sheet(wb, hits: list) -> str:
    """検出結果シートを新規作成し、一覧（リンク付き）を書き込む。作ったシート名を返す。"""
    title = unique_sheet_title(wb.sheetnames)
    ws = wb.create_sheet(title)

    headers = ["No.", "シート名", "セル番地（クリックで移動）", "セルの値"]
    ws.append(headers)
    for i, (sheet_name, coord, value) in enumerate(hits, start=1):
        link = build_link_formula(sheet_name, coord, coord)
        ws.append([i, sheet_name, link, str(value)])
        # 値の文字が "=..." だと Excel に数式と誤解される。最終列を「文字列」型に固定して防ぐ
        ws.cell(row=ws.max_row, column=4).data_type = "s"

    # 見やすさのための軽い整え（幅・1行目固定・見出しを太字）
    widths = [6, 20, 24, 50]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"

    # 検出結果シートを先頭に持ってくる（開いてすぐ見えるように）
    wb.move_sheet(title, -(len(wb.sheetnames) - 1))
    ws.sheet_view.tabSelected = True
    wb.active = wb.sheetnames.index(title)
    return title


# ===== 画面とのやり取り =====

def ask(prompt: str, default: str = "") -> str:
    suffix = f"（未入力なら {default}）" if default else ""
    value = input(f"{prompt}{suffix}: ").strip()
    return value if value else default


def pause_and_exit():
    try:
        input("\nEnterキーで終了します…")
    except EOFError:
        pass


def main():
    print("=" * 50)
    print(" Excel『NG箇所』検出＆一覧化ツール（Day 026）")
    print("=" * 50)

    raw = ask("調べたいExcelファイルのパス（.xlsx）", "")
    if not raw:
        print("⚠ ファイルのパスが入力されませんでした。")
        pause_and_exit()
        return

    path = Path(raw.strip('"'))  # ドラッグ＆ドロップで付く " を外す
    if not path.is_file():
        print(f"⚠ ファイルが見つかりません: {path}")
        pause_and_exit()
        return
    if path.suffix.lower() != ".xlsx":
        print("⚠ このツールは .xlsx 専用です（古い .xls は未対応）。Excelで .xlsx 形式に保存し直してください。")
        pause_and_exit()
        return

    keyword = ask("探す文字（例: NG）", "NG")
    if not keyword:
        print("⚠ 探す文字が空です。")
        pause_and_exit()
        return

    exact = ask("一致のしかた: 1) 含む（部分一致）  2) ぴったり同じ（完全一致）", "1")
    mode = "exact" if exact.strip() == "2" else "contains"
    case_sensitive = ask("大文字小文字を区別しますか？ (yes/no)", "no").lower() in ("yes", "y")

    # ブックを2通りで開く:
    #   wb_save   … 結果シートを足して別名保存する用（数式はそのまま）
    #   wb_values … 検索する用。data_only=True で「数式の計算結果」を見る
    #               （=IF(...,"NG",...) のような式の“本文”を誤って拾わないため）
    try:
        wb_save = openpyxl.load_workbook(path)
        wb_values = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"⚠ Excelファイルを開けませんでした（壊れている／パスワード付き／対応外の可能性）: {e}")
        pause_and_exit()
        return

    # 過去にこのツールで作った「検出結果」シートは調べ直さない（誤検出・二重集計を防ぐ）
    skip = {t for t in wb_values.sheetnames if t == "検出結果" or t.startswith("検出結果(")}
    if skip:
        print(f"（検索から除外したシート: {', '.join(sorted(skip))}）")
    hits = find_matches(wb_values, keyword, mode, case_sensitive, skip)

    mode_label = "完全一致" if mode == "exact" else "部分一致"
    print(f"\n『{keyword}』を {mode_label}で検索 … {len(hits)} 件見つかりました。")
    if not hits:
        print("該当なし。新しいファイルは作りませんでした。")
        pause_and_exit()
        return

    # ざっと中身を確認（多い時は先頭10件だけ表示）
    for sheet_name, coord, value in hits[:10]:
        print(f"   {sheet_name} ! {coord} : {value}")
    if len(hits) > 10:
        print(f"   … ほか {len(hits) - 10} 件")

    # 元ファイルと同じ場所に別名（〇〇_検出結果.xlsx）で保存。シート作成〜保存をまとめて守る
    out = unique_path(path.parent / f"{path.stem}_検出結果.xlsx")
    try:
        title = write_result_sheet(wb_save, hits)
        wb_save.save(out)
    except OSError as e:
        print(f"⚠ 保存に失敗しました（出力先のファイルをExcelで開いたままにしていませんか？）: {e}")
        pause_and_exit()
        return
    except Exception as e:
        print(f"⚠ 結果ファイルの作成に失敗しました: {e}")
        pause_and_exit()
        return

    print(f"\n✅ 完了しました。一覧シート「{title}」を作って保存しました。")
    print(f"   保存先: {out}")
    print("   一覧のセル番地をクリックすると、元の箇所へジャンプできます。")
    pause_and_exit()


if __name__ == "__main__":
    main()
